"""
Genera data/historico_normalizado.xlsx a partir de data/historico.csv.

El XLSX tiene 3 hojas, estilizadas para presentación al equipo de Plaza Vea:
  1. Datos Normalizados     — el dataset completo limpio, listo para Power BI
  2. Resumen Competitivo    — agregación por supermercado × categoría
  3. Diccionario de Campos  — documentación de cada columna

Uso:
    python build_excel.py                         # lee data/historico.csv
    python build_excel.py --input otra/ruta.csv   # input custom
    python build_excel.py --output otro.xlsx      # output custom

El script ES llamado automáticamente al final de scrape_all.py, pero también
se puede ejecutar standalone si solo querés regenerar el XLSX.
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from normalizar import normalizar_df, construir_resumen


# ── Estilos ─────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E5090"
GRAY_ROW   = "F5F7FA"
WHITE      = "FFFFFF"

font_title    = Font(name="Arial", bold=True, color=WHITE, size=13)
font_header   = Font(name="Arial", bold=True, color=WHITE, size=10)
font_data     = Font(name="Arial", size=9)
font_desc_si  = Font(name="Arial", size=9, bold=True, color="1E6B2B")

fill_dark    = PatternFill("solid", start_color=DARK_BLUE)
fill_mid     = PatternFill("solid", start_color=MID_BLUE)
fill_gray    = PatternFill("solid", start_color=GRAY_ROW)
fill_white   = PatternFill("solid", start_color=WHITE)

thick_bottom = Border(bottom=Side(style="medium", color=MID_BLUE))
hair_bottom  = Border(bottom=Side(style="hair",   color="DEDEDE"))

align_center = Alignment(horizontal="center", vertical="center")
align_left   = Alignment(horizontal="left",   vertical="center")
align_right  = Alignment(horizontal="right",  vertical="center")

FMT_CURRENCY = "#,##0.00"
FMT_PCT      = "0.00"
FMT_INT      = "#,##0"


# ── Anchos de columna ───────────────────────────────────────────────────────
ANCHOS_DATOS = {
    "Supermercado": 14, "Categoría": 30, "Familia": 25,
    "Nombre Producto": 45, "ID Producto": 14, "SKU": 14, "Marca": 20,
    "Precio Efectivo (S/)": 16,
    "Precio Tarjeta (S/)": 16, "Nombre Tarjeta": 16, "Descuento Tarjeta %": 17,
    "Precio Internet (S/)": 16, "Precio Normal (S/)": 16,
    "Precio Descuento (S/)": 18, "Precio Regular (S/)": 16,
    "Tiene Descuento": 14, "Descuento %": 12, "Fecha Extracción": 16,
    "URL": 50, "URL Imagen": 50, "Vendedor": 22,
}

ANCHOS_RESUMEN = {
    "Supermercado": 14, "Categoría": 35, "Total Productos": 16,
    "Con Precio": 12, "Precio Mín (S/)": 14, "Precio Máx (S/)": 14,
    "Precio Promedio (S/)": 18, "Con Descuento": 14, "Descuento Promedio %": 20,
}


# ── Helpers de escritura ────────────────────────────────────────────────────

def escribir_banner(ws, texto, n_cols):
    """Fila 1: banner azul con título."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=texto)
    c.font = font_title
    c.fill = fill_dark
    c.alignment = align_center
    ws.row_dimensions[1].height = 26


def escribir_headers(ws, cols, fila=2):
    """Fila de encabezados de tabla."""
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=fila, column=i, value=name)
        c.font = font_header
        c.fill = fill_mid
        c.alignment = align_center
        c.border = thick_bottom
    ws.row_dimensions[fila].height = 22


def aplicar_formato_celda(cell, col_name, value):
    """Aplica formato y alineación según el tipo de columna."""
    cell.font = font_data
    cell.border = hair_bottom
    if "(S/)" in col_name:
        cell.number_format = FMT_CURRENCY
        cell.alignment = align_right
    elif "%" in col_name:
        cell.number_format = FMT_PCT
        cell.alignment = align_right
    elif col_name == "Tiene Descuento":
        cell.alignment = align_center
        if value == "Sí":
            cell.font = font_desc_si
    elif col_name in ("ID Producto", "SKU", "Fecha Extracción"):
        cell.alignment = align_center
    elif col_name in ("Total Productos", "Con Precio", "Con Descuento"):
        cell.number_format = FMT_INT
        cell.alignment = align_right
    elif col_name == "Nombre Tarjeta":
        cell.alignment = align_center
    else:
        cell.alignment = align_left


def escribir_data(ws, df, start_row=3):
    """Escribe el DataFrame con zebra striping."""
    cols = list(df.columns)
    for r_offset, row in enumerate(df.itertuples(index=False)):
        r = start_row + r_offset
        fill = fill_gray if r % 2 == 0 else fill_white
        for c_idx, value in enumerate(row, start=1):
            # Manejar NaN/None para openpyxl
            v = value if pd.notna(value) else None
            cell = ws.cell(row=r, column=c_idx, value=v)
            cell.fill = fill
            aplicar_formato_celda(cell, cols[c_idx-1], v)


def aplicar_anchos(ws, cols, anchos):
    for i, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(name, 18)


# ── Construcción del workbook ───────────────────────────────────────────────

def construir_workbook(df_norm: pd.DataFrame, resumen: pd.DataFrame,
                       fecha_extraccion: str) -> Workbook:
    wb = Workbook()

    # Hoja 1 — Datos
    ws1 = wb.active
    ws1.title = "Datos Normalizados"
    ws1.freeze_panes = "A3"
    escribir_banner(
        ws1,
        f"📊 Comparativo de Precios – Plaza Vea / Tottus / Wong / Metro  "
        f"|  Extracción: {fecha_extraccion}",
        n_cols=len(df_norm.columns),
    )
    escribir_headers(ws1, df_norm.columns)
    escribir_data(ws1, df_norm)
    aplicar_anchos(ws1, df_norm.columns, ANCHOS_DATOS)

    # Hoja 2 — Resumen
    ws2 = wb.create_sheet("Resumen Competitivo")
    ws2.freeze_panes = "A3"
    escribir_banner(
        ws2,
        "📈 Resumen Competitivo por Supermercado y Categoría",
        n_cols=len(resumen.columns),
    )
    escribir_headers(ws2, resumen.columns)
    escribir_data(ws2, resumen)
    aplicar_anchos(ws2, resumen.columns, ANCHOS_RESUMEN)

    # Hoja 3 — Diccionario
    ws3 = wb.create_sheet("Diccionario de Campos")
    escribir_banner(ws3, "Diccionario de Campos – Datos Normalizados", n_cols=3)
    escribir_headers(ws3, ["Campo", "Tipo", "Descripción"])
    diccionario = [
        ("Supermercado",          "Texto",  "Nombre del supermercado: Plaza Vea, Tottus, Wong, Metro"),
        ("Categoría",             "Texto",  "Categoría del producto (ej. Carne De Res, Carne De Pollo)"),
        ("Familia",               "Texto",  "Familia agrupadora cross-supermercado (res, pollo, frutas, etc.)"),
        ("Nombre Producto",       "Texto",  "Nombre descriptivo tal como aparece en la web"),
        ("ID Producto",           "Texto",  "Identificador único del producto en el sistema del supermercado"),
        ("SKU",                   "Texto",  "Código SKU específico de la variante"),
        ("Marca",                 "Texto",  "Marca declarada por el supermercado (sub-marcas se mantienen separadas)"),
        ("Precio Efectivo (S/)",  "Número", "Mejor precio para el cliente. Prioridad: Tarjeta > Descuento > Internet > Normal > Regular"),
        ("Precio Tarjeta (S/)",   "Número", "Precio pagando con la tarjeta del supermercado. Tottus: exacto (del CMR). Plaza Vea: APROXIMADO (calculado desde el % de la promoción Oh!, puede diferir ligeramente del precio real en checkout). Wong y Metro: no aplica."),
        ("Nombre Tarjeta",        "Texto",  "Tarjeta a la que aplica el precio: 'CMR' (Tottus) o 'Tarjeta Oh!' (Plaza Vea). Vacío si no aplica."),
        ("Descuento Tarjeta %",   "Número", "Porcentaje de descuento exclusivo con la tarjeta del supermercado"),
        ("Precio Internet (S/)",  "Número", "Precio web sin tarjeta. Es el precio que paga el cliente sin promoción."),
        ("Precio Normal (S/)",    "Número", "Precio referencial pre-promoción. Vacío si no aplica."),
        ("Precio Descuento (S/)", "Número", "Precio con descuento aplicado. Vacío si no aplica."),
        ("Precio Regular (S/)",   "Número", "Precio sin descuentos. Vacío si no aplica."),
        ("Tiene Descuento",       "Texto",  "Indicador 'Sí' / 'No' (descuento online, no incluye tarjeta)"),
        ("Descuento %",           "Número", "Porcentaje de descuento online (0 si no hay)"),
        ("Fecha Extracción",      "Fecha",  "Fecha del scraping (YYYY-MM-DD)"),
        ("URL",                   "Texto",  "Link al producto en la web del supermercado"),
        ("URL Imagen",            "Texto",  "Link a la imagen del producto"),
        ("Vendedor",              "Texto",  "Vendedor registrado en la plataforma"),
    ]
    for r_offset, (campo, tipo, desc) in enumerate(diccionario):
        r = 3 + r_offset
        fill = fill_gray if r % 2 == 0 else fill_white
        for c_idx, val in enumerate([campo, tipo, desc], start=1):
            cell = ws3.cell(row=r, column=c_idx, value=val)
            cell.font = font_data
            cell.fill = fill
            cell.border = hair_bottom
            cell.alignment = align_left
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 75

    return wb


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  default="data/historico.csv",
                        help="CSV de entrada (default: data/historico.csv)")
    parser.add_argument("--output", default="data/historico_normalizado.xlsx",
                        help="XLSX de salida (default: data/historico_normalizado.xlsx)")
    args = parser.parse_args()

    in_path  = Path(args.input)
    out_path = Path(args.output)

    if not in_path.exists():
        print(f"✗ No se encontró {in_path}")
        print("  Asegurate de haber corrido scrape_all.py al menos una vez.")
        sys.exit(1)

    print(f"📖 Leyendo {in_path}...")
    df_raw = pd.read_csv(in_path, dtype=str, encoding="utf-8-sig")
    print(f"   {len(df_raw):,} filas, {len(df_raw.columns)} columnas")

    print("🧹 Normalizando...")
    df_norm = normalizar_df(df_raw)

    print("📊 Construyendo resumen...")
    resumen = construir_resumen(df_norm)
    print(f"   {len(resumen)} grupos (supermercado × categoría)")

    fecha = datetime.now().strftime("%Y-%m-%d")
    print(f"📝 Generando XLSX...")
    wb = construir_workbook(df_norm, resumen, fecha)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"✅ Guardado: {out_path}")
    print(f"   Hoja 1: Datos Normalizados   ({len(df_norm):,} filas)")
    print(f"   Hoja 2: Resumen Competitivo  ({len(resumen)} filas)")
    print(f"   Hoja 3: Diccionario de Campos")


if __name__ == "__main__":
    main()
